import wx
import wx.lib
import wx.lib.scrolledpanel as scrolled
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os
import json
import re
from twilio.rest import Client
from openpyxl import load_workbook


class SelectGUI(wx.Dialog):
    def __init__(self, parent, title, recipients_list, sid, authtoken, phonenumber):

        list_size = len(recipients_list)
        self.ui_size = (400, list_size*25 + 100)
        if self.ui_size[1] > 260:
            self.ui_size = (400, 260)

        super(SelectGUI, self).__init__(parent, title=title, size=self.ui_size)
        self.Center()

        self.recipients_list = recipients_list
        self.sid = sid
        self.authtoken = authtoken
        self.phonenumber = phonenumber
        self.receipients_checkbox = []

        self.init_ui()

    def init_ui(self):
        # Set font for the GUI Dialog
        font = wx.SystemSettings.GetFont(wx.SYS_DEFAULT_GUI_FONT)
        font.SetPointSize(10)

        #Scroll Panel Containing All possible recipients
        checkbox_panel = scrolled.ScrolledPanel(self, -1, size=(self.ui_size[0]-15, self.ui_size[1]-75), pos=(0, 0), style=wx.SIMPLE_BORDER)
        checkbox_panel.SetBackgroundColour(wx.Colour(red=245, blue=245, green=245))
        scroll_vbox = wx.BoxSizer(wx.VERTICAL)

        # Now that the scroll panel has been created, we can create our list of checkbox items
        self.receipients_checkbox = self.create_checkbox_list(self.recipients_list, checkbox_panel)

        scroll_vbox.Add(self.receipients_checkbox[0], flag=wx.TOP | wx.LEFT, border=5)
        scroll_vbox.Add(wx.StaticLine(checkbox_panel, -1, size=(365, -1)), 0, wx.TOP, 2)

        for cb in self.receipients_checkbox[1:len(self.receipients_checkbox)]:
            scroll_vbox.Add(cb, flag=wx.TOP | wx.LEFT, border=5)

        checkbox_panel.SetSizer(scroll_vbox)
        # Call this after adding all of the objects
        checkbox_panel.SetupScrolling()

        # Create panel to hold button that sends message or cancels
        buttons_panel = wx.Panel(self, style=wx.SIMPLE_BORDER, size=(400, 37), pos=(0, self.ui_size[1]-75))
        buttons_panel.SetBackgroundColour(wx.Colour(red=255, blue=255, green=255))

        buttons_hbox = wx.BoxSizer(wx.HORIZONTAL)

        info_label = wx.StaticText(buttons_panel, label='Select all recipients to receive message', pos=(0, 7))
        info_label.SetFont(font)
        buttons_hbox.Add(info_label)
        accept_button = wx.Button(buttons_panel, label='Accept', size=(50, 30), pos=(275, 2))
        buttons_hbox.Add(accept_button, flag=wx.RIGHT, border=0)
        cancel_button = wx.Button(buttons_panel, label='Cancel', size=(50, 30), pos=(330, 2))
        buttons_hbox.Add(cancel_button, flag=wx.LEFT | wx.BOTTOM | wx.ALIGN_RIGHT, border=0)

        buttons_panel.SetSizer(buttons_hbox)

        # Bind buttons to event handler
        cancel_button.Bind(wx.EVT_BUTTON, self.on_cancel)
        accept_button.Bind(wx.EVT_BUTTON, self.on_accept)

    def create_checkbox_list(self, recipients_list, parent):
        temp_list = []
        all_cb = wx.CheckBox(parent, label="Select All")
        all_cb.Bind(wx.EVT_CHECKBOX, self.on_all_check)
        temp_list.append(all_cb)
        for member in recipients_list:
            cb = wx.CheckBox(parent, label=member.first_name + " " + member.last_name)
            cb.Bind(wx.EVT_CHECKBOX, self.on_cb_check)
            temp_list.append(cb)
        return temp_list

    # Select all of the options if the Select All Checkbox is checked, remove if unchecked
    def on_all_check(self, event):
        all_cb = event.GetEventObject()
        is_checked = all_cb.GetValue()
        if is_checked:
            for cb in self.receipients_checkbox:
                cb.SetValue(1)
        else:
            for cb in self.receipients_checkbox:
                cb.SetValue(0)

    # If a box is unchecked and the Select All Checkbox is checked, remove the select all check
    def on_cb_check(self, event):
        cb = event.GetEventObject()
        is_checked = cb.GetValue()
        if not is_checked:
            self.receipients_checkbox[0].SetValue(0)

    def on_cancel(self, event):
        self.EndModal(0)

    def on_accept(self, event):
        self.EndModal(1)


class ConfigGUI(wx.Dialog):
    def __init__(self, parent, title, config_data_json, config_file):
        super(ConfigGUI, self).__init__(parent, title=title, size=(400, 185))

        self.config_file = config_file
        self.config_data_json = config_data_json
        self.sid_tc = wx.TextCtrl()
        self.authtoken_tc = wx.TextCtrl()
        self.phonenumber_tc = wx.TextCtrl()

        self.init_ui()
        self.Bind(wx.EVT_CLOSE, self.on_cancel)

        self.get_config_data()

        self.Center()

        self.Show()

    def init_ui(self):
        font = wx.SystemSettings.GetFont(wx.SYS_DEFAULT_GUI_FONT)
        font.SetPointSize(10)

        panel = wx.Panel(self, style=wx.SIMPLE_BORDER)
        panel.SetBackgroundColour(wx.Colour(red=245, blue=245, green=245))

        # Main vertical sizer box
        vertical_sizer_box = wx.BoxSizer(wx.VERTICAL)

        sid_hbox = wx.BoxSizer(wx.HORIZONTAL)
        sid_stext = wx.StaticText(panel, label='Account SID:')
        sid_stext.SetFont(font)
        sid_hbox.Add(sid_stext, flag=wx.ALIGN_CENTER_VERTICAL | wx.LEFT | wx.TOP, border=5)
        sid_hbox.Add((19, 0))
        self.sid_tc = wx.TextCtrl(panel, size=(250, 22))
        sid_hbox.Add(self.sid_tc, proportion=1, flag=wx.ALIGN_CENTER_VERTICAL | wx.EXPAND | wx.TOP | wx.LEFT,
                     border=5)
        vertical_sizer_box.Add(sid_hbox, flag=wx.RIGHT | wx.EXPAND, border=10)

        # Add gap
        vertical_sizer_box.Add((-1, 10))

        authtoken_hbox = wx.BoxSizer(wx.HORIZONTAL)
        authtoken_stext = wx.StaticText(panel, label='Auth Token:')
        authtoken_stext.SetFont(font)
        authtoken_hbox.Add(authtoken_stext, flag=wx.ALIGN_CENTER_VERTICAL | wx.LEFT, border=5)
        authtoken_hbox.Add((23, 0))
        self.authtoken_tc = wx.TextCtrl(panel, size=(250, 22))
        authtoken_hbox.Add(self.authtoken_tc, proportion=1, flag=wx.ALIGN_CENTER_VERTICAL | wx.EXPAND | wx.LEFT,
                     border=5)
        vertical_sizer_box.Add(authtoken_hbox, flag=wx.RIGHT | wx.EXPAND, border=10)

        # Add gap
        vertical_sizer_box.Add((-1, 10))

        phonenumber_hbox = wx.BoxSizer(wx.HORIZONTAL)
        phonenumber_stext = wx.StaticText(panel, label='Phone Number:')
        phonenumber_stext.SetFont(font)
        phonenumber_hbox.Add(phonenumber_stext, flag=wx.ALIGN_CENTER_VERTICAL | wx.LEFT, border=5)
        self.phonenumber_tc = wx.TextCtrl(panel, size=(250, 22))
        phonenumber_hbox.Add(self.phonenumber_tc, proportion=1, flag=wx.ALIGN_CENTER_VERTICAL | wx.EXPAND | wx.LEFT,
                     border=5)
        vertical_sizer_box.Add(phonenumber_hbox, flag=wx.RIGHT | wx.EXPAND, border=10)

        # Adding gap before buttons
        vertical_sizer_box.Add((-1, 15))

        buttons_hbox = wx.BoxSizer(wx.HORIZONTAL)
        buttons_hbox.Add((0, 10), 1, flag=wx.EXPAND)
        ok_button = wx.Button(panel, label='OK', size=(50, 30))
        buttons_hbox.Add(ok_button, flag=wx.RIGHT, border=5)
        cancel_button = wx.Button(panel, label='Cancel', size=(50, 30))
        buttons_hbox.Add(cancel_button, flag=wx.LEFT | wx.BOTTOM | wx.ALIGN_RIGHT, border=0)
        vertical_sizer_box.Add(buttons_hbox, flag=wx.RIGHT | wx.EXPAND, border=10)

        panel.SetSizer(vertical_sizer_box)

        self.Bind(wx.EVT_BUTTON, self.on_ok, id=ok_button.GetId())
        self.Bind(wx.EVT_BUTTON, self.on_cancel, id=cancel_button.GetId())

    def get_config_data(self):
        self.sid_tc.SetValue(self.config_data_json.get("sid"))
        self.authtoken_tc.SetValue(self.config_data_json.get("authtoken"))
        self.phonenumber_tc.SetValue(self.config_data_json.get("phonenumber"))

    def on_ok(self, event):
        sid = self.sid_tc.GetValue()
        authtoken = self.authtoken_tc.GetValue()
        phonenumber = self.phonenumber_tc.GetValue()

        self.config_data_json['sid'] = sid
        self.config_data_json['authtoken'] = authtoken
        self.config_data_json['phonenumber'] = phonenumber
        self.config_file.truncate(0)
        self.config_file.seek(0, 0)
        string = json.dumps(self.config_data_json)
        self.config_file.write(string)
        self.config_file.flush()
        self.Destroy()

    def on_cancel(self, event):
        self.Destroy()


class PrimaryGUI(wx.Frame):
    def __init__(self, parent, title, sid, authtoken, phonenumber, default_recipient_file, config_data_json, config_file):
        super(PrimaryGUI, self).__init__(parent, title=title, size=(700, 400))

        # Initialize dynamic variables

        self.config_file = config_file
        self.config_data_json = config_data_json
        self.recipients_file_location = default_recipient_file
        self.sid_value = sid
        self.auth_token_value = authtoken
        self.phonenumber_value = phonenumber

        # Initialize a dynamic TextCtrl field for the message and recipients file
        self.message_textctrl = wx.TextCtrl()
        self.recipients_tc = wx.TextCtrl()

        self.Bind(wx.EVT_CLOSE, self.on_close)

        # Initialize UI Functions
        self.init_ui()

        # Center the screen
        self.Centre()

        # Set the minimum screen size
        self.SetMinSize(size=(650, 400))

        # Make the screen visible to the user
        self.Show()

    #still need to store the most recent receptiants file here
    def on_close(self, event):
        self.Destroy()

    # Initialize the UI
    def init_ui(self):

        # Get system font
        font = wx.SystemSettings.GetFont(wx.SYS_DEFAULT_GUI_FONT)
        font.SetPointSize(10)

        # Set up the menubar
        menubar = wx.MenuBar()
        configuration_menu = wx.Menu()
        menubar.Append(configuration_menu, '&Configuration')
        self.SetMenuBar(menubar)

        panel = wx.Panel(self, style=wx.SIMPLE_BORDER)
        panel.SetBackgroundColour(wx.Colour(red=245, blue=245, green=245))

        # Main vertical sizer box
        vertical_sizer_box = wx.BoxSizer(wx.VERTICAL)

        # Adds Enter Message dialog
        label_hbox = wx.BoxSizer(wx.HORIZONTAL)
        label_stext = wx.StaticText(panel, label='Enter Message:')
        label_stext.SetFont(font)
        label_hbox.Add(label_stext)
        vertical_sizer_box.Add(label_hbox, flag=wx.LEFT | wx.TOP, border=5)

        # Adds text box to enter dialog
        message_hbox = wx.BoxSizer(wx.HORIZONTAL)
        self.message_textctrl = wx.TextCtrl(panel, style=wx.TE_MULTILINE)
        message_hbox.Add(self.message_textctrl, proportion=1, flag=wx.EXPAND)
        vertical_sizer_box.Add(message_hbox, proportion=1, flag=wx.LEFT | wx.RIGHT | wx.EXPAND | wx.TOP, border=5)

        # Add gap between box and buttons
        vertical_sizer_box.Add((-1, 10))

        # Add buttons and field for recipients file
        buttons_hbox = wx.BoxSizer(wx.HORIZONTAL)
        recipients_stext = wx.StaticText(panel, label='Recipients File:')
        recipients_stext.SetFont(font)
        buttons_hbox.Add(recipients_stext, flag=wx.ALIGN_CENTER_VERTICAL | wx.LEFT, border=5)
        self.recipients_tc = wx.TextCtrl(panel, size=(300, 22))
        self.recipients_tc.SetValue(self.recipients_file_location)
        buttons_hbox.Add(self.recipients_tc, proportion=0, flag=wx.ALIGN_CENTER_VERTICAL | wx.LEFT | wx.RIGHT, border=5)
        select_file_button = wx.Button(panel, label='Select File', size=(70, 30))
        buttons_hbox.Add(select_file_button)
        buttons_hbox.Add((0, 10), 1, flag=wx.EXPAND)
        tags_button = wx.Button(panel, label='Tags', size=(50, 30))
        buttons_hbox.Add(tags_button, flag=wx.RIGHT, border=5)
        send_message_button = wx.Button(panel, label='Send Message', size=(100, 30))
        buttons_hbox.Add(send_message_button, flag=wx.LEFT | wx.BOTTOM | wx.ALIGN_RIGHT, border=0)
        vertical_sizer_box.Add(buttons_hbox, flag=wx.RIGHT | wx.EXPAND, border=10)

        # Add gap a bottom of screen
        vertical_sizer_box.Add((-1, 10))

        panel.SetSizer(vertical_sizer_box)

        # Add events to buttons
        self.Bind(wx.EVT_BUTTON, self.on_send_message, id=send_message_button.GetId())
        self.Bind(wx.EVT_BUTTON, self.on_select_file, id=select_file_button.GetId())
        self.Bind(wx.EVT_BUTTON, self.on_tags, id=tags_button.GetId())

        # This event is called any time a menu is about to be opened
        # This should be changed to work only on the configuration menu if more menus are added in future versions
        self.Bind(wx.EVT_MENU_OPEN, self.on_configuration)

    # Load all of the recipients from the selected excel spreadsheet
    def load_recipients(self):
        wb = load_workbook(self.recipients_file_location)
        sheet = wb.worksheets[0]
        recipients_list = []

        if (sheet.cell(row=1, column=1).value is None or "bond" not in sheet.cell(row=1, column=1).value.lower() or
            sheet.cell(row=1, column=2).value is None or "first" not in sheet.cell(row=1, column=2).value.lower() or
            sheet.cell(row=1, column=3).value is None or "last" not in sheet.cell(row=1, column=3).value.lower() or
            sheet.cell(row=1, column=4).value is None or "phone" not in sheet.cell(row=1, column=4).value.lower()):
            return None


        for i in range(2, sheet.max_row + 1):
            recipients_list.append(Recipient(bond=sheet.cell(row=i, column=1).value,
                                             first_name=sheet.cell(row=i, column=2).value,
                                             last_name=sheet.cell(row=i, column=3).value,
                                             phonenumber=self.format_phone_number(sheet.cell(row=i, column=4).value)))

        # Return the list of recipient objects
        return recipients_list

    # Here we will put the code for when the message is going to be sent
    def on_send_message(self, event):
        msg = self.message_textctrl.GetValue()
        if msg is "":
            wx.MessageBox('Cannot send empty message!', 'Error', wx.OK | wx.ICON_ERROR)
            return

        if self.recipients_file_location is "":
            wx.MessageBox('No file selected!', 'Error', wx.OK | wx.ICON_ERROR)
            return

        try:
            recipients_list = self.load_recipients()
            if recipients_list is None:
                raise FileNotFoundError()
        except FileNotFoundError:
            wx.MessageBox('Error loading file!', 'Error', wx.OK | wx.ICON_ERROR)
            return

        select_ui = SelectGUI(self, "Select Recipients", recipients_list, self.sid_value, self.auth_token_value, self.phonenumber_value)
        return_msg = select_ui.ShowModal()
        if return_msg is 0:
            return
        checkbox_items = select_ui.receipients_checkbox
        select_ui.Destroy()

        response = wx.MessageBox('Are you sure you would like to send this message?', 'Send Message...',
                                 wx.YES_NO | wx.ICON_QUESTION)
        if response is not 2:
            return

        # This error handling needs some work
        try:
            twilioCli = Client(self.sid_value, self.auth_token_value)
            for i in range(1, len(recipients_list)+1):
                if checkbox_items[i].GetValue():
                    format_msg = self.tag_matching(msg, recipients_list[i-1])
                    twilioCli.messages.create(body=format_msg, from_=self.phonenumber_value, to=recipients_list[i-1].phonenumber)
                    #print(recipients_list[i-1].phonenumber)
        # This is too broad, but I'm not sure how to handle twilio exceptions right now
        except:
            wx.MessageBox('Error: Check Twilio configuration!', 'Error', wx.OK | wx.ICON_ERROR)
            return

        wx.MessageBox('Your message was sent successfully!', 'Sent', wx.OK | wx.ICON_INFORMATION)

    def tag_matching(self, msg, recipient):
        rep = {"#Bond": str(recipient.bond),
                "#FirstName": str(recipient.first_name),
                "#LastName": str(recipient.last_name),
                "#PhoneNumber": str(recipient.phonenumber),
               }
        rep = dict((re.escape(k), v) for k, v in rep.items())
        pattern = re.compile("|".join(rep.keys()))
        return pattern.sub(lambda m: rep[re.escape(m.group(0))], msg)

    def format_phone_number(self, number):
        number = number.replace("(", '')
        number = number.replace(")", '')
        number = number.replace("-", '')
        number = "+1" + number
        return number

    # Using tKinter to allow the user to select a file location
    def on_select_file(self, event):

        # Open up the GUI for selecting a file
        Tk().withdraw()  # we don't want a full GUI, so keep the root window from appearing
        filename = askopenfilename(title="Choose a File",
                                   initialdir=os.path.dirname(os.path.realpath(__file__)),
                                   filetypes=(("Excel File (*.xlsx)", "*.xlsx"),
                                              ("Legacy Excel File (*.xls)", "*.xls"),
                                              ),
                                   )  # show an "Open" dialog box and return the path to the selected file

        # If  a file was selected, store it and put it on the screen
        if filename is not "":
            self.recipients_file_location = filename
            self.recipients_tc.SetValue(filename)
            self.config_data_json['default_recipient_file'] = self.recipients_file_location
            self.config_file.truncate(0)
            self.config_file.seek(0, 0)
            string = json.dumps(self.config_data_json)
            self.config_file.write(string)
            self.config_file.flush()

    # Open a message dialog showing the current tag options
    def on_tags(self, event):
        dial = wx.MessageDialog(None, 'Check Documentation for Information on Tags\n'
                                      '#FirstName\n'
                                      '#LastName\n'
                                      '#Bond\n'
                                      '#PhoneNumber'
                                      , 'Tags', wx.OK | wx.ICON_QUESTION)# This isn't actually a question
                                                                                       # But theres an annoying sound
                                                                                       # otherwise
        dial.ShowModal()

    # Here we will show the configuration menu
    def on_configuration(self, event):
        config_ui = ConfigGUI(self, "Configuration", self.config_data_json, self.config_file)
        config_ui.ShowModal()
        self.config_data_json = config_ui.config_data_json
        self.sid_value = self.config_data_json.get('sid')
        self.auth_token_value = self.config_data_json.get('authtoken')
        self.phonenumber_value = self.config_data_json.get('phonenumber')


class Recipient():
    def __init__(self, bond="", first_name="", last_name="", phonenumber=""):
        self.bond = bond
        self.first_name = first_name
        self.last_name = last_name
        self.phonenumber = phonenumber
        self.checked = False
